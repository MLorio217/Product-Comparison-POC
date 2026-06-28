import re
import io
import tempfile
from pathlib import Path

import fitz
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font
from difflib import SequenceMatcher

st.set_page_config(
    page_title="Product Comparison Generator",
    page_icon="📊",
    layout="centered"
)

# ─────────────────────────────────────────────
# PDF PARSER
# ─────────────────────────────────────────────
NOISE_TERMS = [
    "fictional", "parser testing", "proposal for",
    "accident hospital care", "accident care", "common injuries",
    "dislocations closed", "fractures closed", "additional benefits",
    "monthly rates", "closed reduction", "open reduction",
    "testing notice", "plan highlights", "human review",
    "administrative notes", "testing notes", "proposal id",
    "effective date", "prepared for", "guaranteed issue",
    "portable coverage", "employee-paid", "rate guarantee",
    "this document", "not intended",
    "summit insurance solutions", "apex benefits group",
    "demo notice", "does not represent", "entirely fictional",
]
STANDALONE_SKIP = {"benefit", "amount", "benefit amount", "employee benefit", "employee"}

def _clean(line):
    return re.sub(r"\s+", " ", str(line).strip())

def _is_noise(line):
    low = line.lower().strip()
    if low in STANDALONE_SKIP:
        return True
    return any(t in low for t in NOISE_TERMS) or len(line) < 2

def _looks_like_amount(text):
    return bool(re.match(r"^\$?\d+(,\d{3})*(\.\d+)?(%|/\w+)?$", text.strip()))

def _parse_amount(text):
    text = text.strip().replace("$", "").replace(",", "")
    text = re.sub(r"/\w+$", "", text)
    if "%" in text:
        return text + "%"
    try:
        val = float(text)
        return int(val) if val == int(val) else val
    except ValueError:
        return None

def extract_benefits(pdf_bytes):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    all_lines = []
    for page in doc:
        for raw in page.get_text().splitlines():
            line = _clean(raw)
            if line and not _is_noise(line):
                all_lines.append(line)
    doc.close()
    benefits = {}
    i = 0
    while i < len(all_lines):
        current = all_lines[i]
        if i + 1 < len(all_lines):
            nxt = all_lines[i + 1]
            if _looks_like_amount(nxt):
                val = _parse_amount(nxt)
                if val is not None and len(current) > 3:
                    benefits[current] = val
                i += 2
                continue
        m = re.match(r"^(.+?)\s+(\$?\d+(?:,\d{3})*(?:\.\d+)?(?:%|/\w+)?)$", current)
        if m:
            name = _clean(m.group(1))
            raw_val = m.group(2).strip()
            if name and not _is_noise(name) and len(name) > 3:
                val = _parse_amount(raw_val)
                if val is not None:
                    benefits[name] = val
            i += 1
            continue
        i += 1
    return benefits

# ─────────────────────────────────────────────
# WORKBOOK SCANNER
# ─────────────────────────────────────────────
SKIP_SECTION_TERMS = [
    "accident hospital care", "accident care", "common injuries",
    "dislocations closed", "fractures closed", "accidental death",
    "accidental dismemberment", "catastrophic accident rider",
    "additional benefits", "monthly rates",
    "closed reduction of dislocation", "closed reduction of fracture",
    "laceration benefits are a total", "common carrier",
    "other accidental death", "home modification", "vehicle modification",
]

D_ONLY_ROWS = {12, 14, 15, 17, 18, 19, 20, 22}

# Rows that are text/plan design fields - never write dollar amounts here
TEXT_ONLY_ROWS = {151, 153, 154, 155}

# Rows that are monthly rates - only write if value looks like a real rate (under $500)
RATE_ROWS = {158, 159, 160, 161, 164, 165, 166}

# Maximum reasonable ACC benefit amount - anything over this is likely a parsing error
MAX_BENEFIT = 200000  # Allows Catastrophic Rider ($100k+) but blocks clearly wrong values

def scan_acc_tab(template_bytes):
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(template_bytes)
    tmp.flush()
    tmp.close()
    wb = load_workbook(tmp.name, data_only=True)
    Path(tmp.name).unlink(missing_ok=True)

    acc_sheet = None
    for name in wb.sheetnames:
        if "acc" in name.lower() and "comparison" in name.lower():
            acc_sheet = name
            break
    if not acc_sheet:
        raise ValueError("Could not find ACC comparison worksheet in template.")

    ws = wb[acc_sheet]
    benefit_map = {}
    for row in ws.iter_rows(min_row=4, max_row=200):
        c_cell = None
        for cell in row:
            if cell.column == 3:
                c_cell = cell
                break
        if not c_cell or not c_cell.value:
            continue
        name = str(c_cell.value).strip()
        if not name or name == ' ':
            continue
        lowered = name.lower()
        if any(skip in lowered for skip in SKIP_SECTION_TERMS):
            continue
        benefit_map[name] = c_cell.row

    return benefit_map, D_ONLY_ROWS, acc_sheet

# ─────────────────────────────────────────────
# DYNAMIC MATCHER
# ─────────────────────────────────────────────
TRANSLATIONS = {
    "initial physician visit":          "initial doctor visit",
    "physician visit":                  "initial doctor visit",
    "urgent care facility":             "urgent care facility treatment",
    "urgent care treatment":            "urgent care facility treatment",
    "urgent care":                      "urgent care facility treatment",
    "emergency room":                   "emergency room treatment",
    "ground emergency transport":       "ground ambulance",
    "air emergency transport":          "air ambulance",
    "icu admission":                    "critical care unit (ccu) admission",
    "icu confinement":                  "critical care unit confinement (per day up to 30 days)",
    "critical care unit admission":     "critical care unit (ccu) admission",
    "open abdominal surgery":           "surgery (open abdominal, thoracic)",
    "exploratory surgery":              "surgery (exploratory or without repair)",
    "outpatient surgery":               "outpatient surgery (once per accident)",
    "physical rehabilitation therapy":  "physical or occupational therapy (per treatment up to 10)",
    "physical therapy":                 "physical or occupational therapy (per treatment up to 10)",
    "rehabilitation therapy":           "physical or occupational therapy (per treatment up to 10)",
    "speech therapy":                   "speech therapy (per treatment up to 10)",
    "mental health therapy":            "mental health therapy (per treatment up to 10)",
    "prescription medication":          "prescription medicine",
    "prescription benefit":             "prescription medicine",
    "major diagnostic imaging":         "major diagnostic exams",
    "advanced imaging":                 "major diagnostic exams",
    "advanced imaging (ct/mri/pet)":    "major diagnostic exams",
    "durable medical equipment":        "medical equipment",
    "blood / plasma / platelets":       "blood, plasma,  platelets",
    "blood / plasma":                   "blood, plasma,  platelets",
    "hospital confinement":             "hospital confinement  (per day up to 365 days)",
    "transportation":                   "transportation (per trip up to 3 per accident)",
    "lodging":                          "lodging (per day up to 30 days)",
    "induced coma":                     "induced coma (up to 14 days)",
    "non-induced coma":                 "non-induced coma (duration of 14 or more days)",
    "rehabilitation facility":          "rehabilitation facility confinement (per day up to 90 days)",
    "physician follow-up visit":        "follow-up doctor treatment",
    "follow up doctor":                 "follow-up doctor treatment",
    "fracture benefit":                 "hip",
    "fracture":                         "hip",
    "laceration benefit":               "laceration (sutures)",
    "burn benefit":                     "burns (2nd degree, at least 36% of body)",
    "burns":                            "burns (2nd degree, at least 36% of body)",
    "accidental death":                 "employee",
    "accident medical expense":         "outpatient surgery (once per accident)",
}

REMOVE_WORDS = ["benefit","benefits","per","day","days","up to","maximum","max","coverage","covered","treatment","level"]

def _normalize(text):
    text = str(text).lower().strip()
    for ch, rep in [("&"," and "),("/"," "),("-"," "),("("," "),(")"," "),("*"," ")]:
        text = text.replace(ch, rep)
    text = re.sub(r"[^a-z0-9 ]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    for word in REMOVE_WORDS:
        text = re.sub(rf"\b{re.escape(word)}\b", " ", text)
    text = re.sub(r"\b\d+\b", " ", text)
    return re.sub(r"\s+", " ", text).strip()

def _translate(name):
    norm = str(name).lower().strip()
    if norm in TRANSLATIONS:
        return TRANSLATIONS[norm]
    for key, val in TRANSLATIONS.items():
        if key in norm:
            return val
    return name

def _similarity(a, b):
    na, nb = _normalize(a), _normalize(b)
    if not na or not nb:
        return 0
    if na == nb:
        return 1.0
    if na in nb or nb in na:
        return 0.92
    seq = SequenceMatcher(None, na, nb).ratio()
    ta, tb = set(na.split()), set(nb.split())
    overlap = len(ta & tb) / max(len(ta), len(tb)) if ta and tb else 0
    return max(seq, overlap)

def find_best_match(pdf_name, benefit_map, threshold=0.62, used_rows=None):
    if used_rows is None:
        used_rows = set()
    translated = _translate(pdf_name)
    for tname, row in benefit_map.items():
        if row in used_rows:
            continue
        if translated.lower().strip() == tname.lower().strip():
            return tname, row, 1.0
    best_name, best_row, best_score = None, None, 0
    search_name = translated if translated != pdf_name else pdf_name
    for tname, row in benefit_map.items():
        if row in used_rows:
            continue
        score = _similarity(search_name, tname)
        if score > best_score:
            best_score = score
            best_name = tname
            best_row = row
    if best_score >= threshold:
        return best_name, best_row, best_score
    return None, None, best_score

# ─────────────────────────────────────────────
# WORKBOOK WRITER
# ─────────────────────────────────────────────
def write_workbook(template_bytes, voya_benefits, competitor_benefits, benefit_map, d_only_rows, review_items):
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(template_bytes)
    tmp.flush()
    tmp.close()
    tmp_path = Path(tmp.name)

    wb = load_workbook(str(tmp_path))
    acc_sheet = None
    for name in wb.sheetnames:
        if "acc" in name.lower() and "comparison" in name.lower():
            acc_sheet = name
            break
    ws = wb[acc_sheet]

    writes = []
    skipped = []

    def write_side(benefits, col):
        used = set()
        for name, value in benefits.items():
            tname, row, score = find_best_match(name, benefit_map, used_rows=used)
            if tname is None:
                skipped.append((name, value, col, f"No match (best score {score:.2f})"))
                continue
            if col == "H" and row in d_only_rows:
                skipped.append((name, value, col, f"Row {row} is D-only"))
                continue
            if row in TEXT_ONLY_ROWS:
                skipped.append((name, value, col, f"Row {row} is a text field - no dollar amount written"))
                continue
            if isinstance(value, (int, float)) and value > MAX_BENEFIT:
                skipped.append((name, value, col, f"Value {value} exceeds max reasonable benefit - likely parsing error"))
                continue
            if row in RATE_ROWS and isinstance(value, (int, float)) and value > 500:
                skipped.append((name, value, col, f"Value {value} too large for rate row {row}"))
                continue
            used.add(row)
            ws[f"{col}{row}"] = value
            writes.append((row, col, name, tname, value, score))

    write_side(voya_benefits, "D")
    write_side(competitor_benefits, "H")

    # Review Needed sheet
    review_name = "Review Needed"
    if review_name in wb.sheetnames:
        ws_r = wb[review_name]
        for row in ws_r.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws_r = wb.create_sheet(review_name)

    headers = ["Source", "PDF Benefit", "PDF Value", "Reason"]
    for col, h in enumerate(headers, 1):
        ws_r.cell(row=1, column=col, value=h).font = Font(bold=True)
    for idx, item in enumerate(review_items, 2):
        ws_r.cell(row=idx, column=1, value=item.get("source", ""))
        ws_r.cell(row=idx, column=2, value=item.get("pdf_benefit", ""))
        ws_r.cell(row=idx, column=3, value=str(item.get("pdf_value", "")))
        ws_r.cell(row=idx, column=4, value=item.get("reason", ""))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    tmp_path.unlink(missing_ok=True)
    return out.getvalue(), writes, skipped

# ─────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────
st.title("📊 Product Comparison Generator")
st.caption("Dynamic mode · Reads benefit rows directly from your template · ACC tab only")
st.markdown("---")

voya_file = st.file_uploader("1. Current / Voya Proposal (PDF)", type=["pdf"])
competitor_file = st.file_uploader("2. Competitor Proposal (PDF)", type=["pdf"])
template_file = st.file_uploader("3. Product Comparison Template (xlsx)", type=["xlsx"])

st.markdown("---")

if st.button("Generate Completed Comparison", type="primary", use_container_width=True):
    if not voya_file:
        st.error("Upload the Current / Voya Proposal PDF.")
    elif not competitor_file:
        st.error("Upload the Competitor Proposal PDF.")
    elif not template_file:
        st.error("Upload the Product Comparison Template.")
    else:
        with st.spinner("Processing..."):
            template_bytes = template_file.read()

            st.write("📋 Scanning template for benefit rows...")
            benefit_map, d_only_rows, acc_sheet = scan_acc_tab(template_bytes)
            st.write(f"✅ Template: {len(benefit_map)} benefit rows found in **{acc_sheet}**")

            st.write("📄 Reading Voya PDF...")
            voya_raw = extract_benefits(voya_file.read())
            st.write(f"✅ Voya: {len(voya_raw)} items extracted")

            st.write("📄 Reading Competitor PDF...")
            comp_raw = extract_benefits(competitor_file.read())
            st.write(f"✅ Competitor: {len(comp_raw)} items extracted")

            review_items = []
            st.write("✍️ Matching and writing values...")

            out_bytes, writes, skipped = write_workbook(
                template_bytes,
                voya_raw,
                comp_raw,
                benefit_map,
                d_only_rows,
                review_items
            )

            # Add skipped to review
            for name, val, col, reason in skipped:
                src = "Voya" if col == "D" else "Competitor"
                review_items.append({"source": src, "pdf_benefit": name, "pdf_value": val, "reason": reason})

            # Re-save with review items
            out_bytes, writes, skipped = write_workbook(
                template_bytes, voya_raw, comp_raw,
                benefit_map, d_only_rows, review_items
            )

            st.write(f"✅ Values written: {len(writes)}")
            if skipped:
                st.write(f"⚠️ Sent to Review Needed: {len(skipped)}")

        st.success("Done! Download your completed workbook below.")
        st.download_button(
            label="⬇️ Download Completed Comparison Workbook",
            data=out_bytes,
            file_name="Completed_Product_Comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
